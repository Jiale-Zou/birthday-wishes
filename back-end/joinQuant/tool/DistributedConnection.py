import redis
import json
import threading
from openpyxl import Workbook
from tool.Tool import default_dump
import time

class RedisQueue:
    ''' 创建一个公共队列，用于储存任务，保留任务结果 '''
    TASK_QUEUE = 'task_queue'
    RESULT_QUEUE = 'result_queue'
    STOP_SIGNAL = 'stop_signal'

    def __init__(self, ip, port, pwd, db=0):
        self.connection = redis.StrictRedis(host=ip, port=port, password=pwd, db=db)

    def clear_database(self, only_used_keys=True):
        ''' 清空数据库 '''
        if only_used_keys:
            self.connection.delete(RedisQueue.TASK_QUEUE)
            self.connection.delete(RedisQueue.RESULT_QUEUE)
            self.connection.delete(RedisQueue.STOP_SIGNAL)
        else:
            for key in self.connection.keys():
                self.connection.delete(key)

    def input_task(self, task):
        ''' 插入任务队列 '''
        if isinstance(task, (int, float, str, bool)) or task is None:
            self.connection.lpush(RedisQueue.TASK_QUEUE, json.dumps(task, default=default_dump))
        elif isinstance(task, (list, tuple, set)):
            pipe = self.connection.pipeline() #使用管道提升批量插入的效率，使得一次网络往返执行多条命令
            for item in task:
                pipe.lpush(RedisQueue.TASK_QUEUE, json.dumps(item, default=default_dump))
            pipe.execute()
        else:
            self.connection.lpush(RedisQueue.TASK_QUEUE, json.dumps(task, default=default_dump))

    def carry_out_task(self, func, other_params=list()):
        '''
        :param func: 要执行的函数
        :param other_param: 执行func函数时，除Redis中获得参数外，额外要输入的参数
        :return: 保存func函数结果到Redis队列中
        '''
        """工作进程主循环"""
        while True:
            # 阻塞式弹出任务，设置超时时间(秒)
            task_data = self.connection.brpop(RedisQueue.TASK_QUEUE, timeout=10)

            # 检测是否收到停止信号
            if self.connection.get(RedisQueue.STOP_SIGNAL):
                print('收到停止信号, 退出！')
                break
            # 如果队列为空，也停止
            if task_data is None:
                print('任务队列为空, 退出！')
                break
            # 解析任务参数（rpop返回的是 b(key,value)）
            _, params_json = task_data
            params = json.loads(params_json)
            if isinstance(params, (str, int, float, bool)):
                params = [params]+other_params
            elif isinstance(params, (list, tuple, set)):
                params = list(params)+other_params
            else:
                params = [params]+other_params
            # 执行计算
            try:
                result = func(*params)
            except:
                result = None #如果计算错误, 则返回None, json.dumps(None)='null', 传到Reids上变为b'null'
                print(f'任务失败: {params}')
            # 存储结果
            self.connection.lpush(RedisQueue.RESULT_QUEUE, json.dumps(result, default=default_dump))

    def signal_handler(self, signum, frame):
        """设置停止信号"""
        self.connection.set(RedisToXLSX.STOP_SIGNAL, '1')


class RedisToXLSX:
    ''' 实时将公共结果队列的数据保存为本地xlsx '''
    TASK_QUEUE = 'task_queue'
    RESULT_QUEUE = 'result_queue'
    STOP_SIGNAL = 'stop_signal'

    def __init__(self, ip, port, pwd, db=0):
        self.connection = redis.StrictRedis(host=ip, port=port, password=pwd, db=db)

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.header_written = False  # 用来标记当前是否有表头
        self.current_row = 1 # 当前写入行

    def clear_database(self, only_used_keys=True):
        ''' 清空数据库 '''
        if only_used_keys:
            self.connection.delete(RedisToXLSX.TASK_QUEUE)
            self.connection.delete(RedisToXLSX.RESULT_QUEUE)
            self.connection.delete(RedisToXLSX.STOP_SIGNAL)
        else:
            for key in self.connection.keys():
                self.connection.delete(key)

    def start_writing(self, path, result_count, header):
        """启动写入线程"""
        writer_thread = threading.Thread(target=self._write_to_xlsx, args=(path, result_count, header))
        writer_thread.daemon = True #守护线程模式（daemon=True）防止主线程退出后线程挂起
        writer_thread.start()
        # 写入线程是daemon，但主线程是非daemon，当主线程没了，会强制子线程结束（不管子线程是否是daemon），故返回子线程，并.join()直至子线程结束
        return writer_thread

    def _write_to_xlsx(self, path, result_count, header):
        '''
        path: 保存的xlsx路径
        result_count: 结果数量，与任务数量一致
        header: xlsx表头
        '''
        cnt = 0 #用来记录当前完成了几个任务

        """持续从Redis获取结果并写入XLSX"""
        while cnt < result_count and not self.connection.get(RedisToXLSX.STOP_SIGNAL):
            # 非阻塞获取结果 (RPOP)
            result_json = self.connection.rpop(RedisToXLSX.RESULT_QUEUE)

            if result_json:
                cnt += 1
                result = json.loads(result_json)
                if result is not None: #该任务没有失败
                    # 如果是第一个结果，写入表头
                    if not self.header_written and isinstance(result, (list, tuple)):
                        for col_num, header in enumerate(header, 1):
                            self.sheet.cell(row=1, column=col_num, value=header)
                        self.header_written = True
                        self.current_row += 1

                    for col_num, value in enumerate(result, 1): # 写入数据行
                        self.sheet.cell(row=self.current_row, column=col_num, value=value)
                    self.current_row += 1
            else:
                # 没有结果时短暂休眠
                time.sleep(0.1)

        self.workbook.save(path) #保存文件
        print('结果保存完毕！')

    def signal_handler(self, signum, frame):
        """ 设置停止信号(用于指定signal.singal) """
        self.connection.set(RedisToXLSX.STOP_SIGNAL, '1')
